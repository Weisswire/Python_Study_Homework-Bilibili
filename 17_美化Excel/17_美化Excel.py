# 通过Python创建excel数据并保存多个数据

from openpyxl import load_workbook,Workbook
from openpyxl.styles import Font,Color,Alignment,PatternFill, colors

infos = [
    ['月份', '销售1组', '销售2组', '销售3组', '销售4组', '销售5组'],
    ['1月', 57, 142, 62, 85, 137],
    ['2月', 149, 114, 125, 113, 100],
    ['3月', 105, 56, 76, 85, 74],
    ['4月', 51, 145, 95, 145, 103],
    ['5月', 116, 113, 118, 102, 136],
    ['6月', 131, 67, 127, 61, 71],
    ['7月', 126, 124, 100, 106, 129],
    ['8月', 148, 113, 117, 113, 121],
    ['9月', 56, 97, 121, 76, 50],
    ['10月', 53, 124, 86, 72, 88],
    ['11月', 130, 63, 131, 61, 124],
    ['12月', 86, 91, 59, 146, 126],
]

def Create_Excel(infos: list):
    wb = Workbook()
    sh = wb.active
    sh.title = '财务报表'
    for info in infos:
        sh.append(info)
    wb.save('财务报表.xlsx')

def Beautify():
    wb = load_workbook('财务报表.xlsx')
    sh = wb.active
    font1 = Font(name='微软雅黑',size=12,bold=True,italic=True,color='66B3FF')
    font2 = Font(name='微软雅黑',size=12,bold=True,italic=False,color='FF60AF')
    font3 = Font(name='微软雅黑',size=12,bold=True,italic=False,color=colors.BLACK)
    for row in sh.rows:
        for cell in row:
            cell.alignment = Alignment(horizontal='center',vertical='center')
    for i in range(1,sh.max_column+1):
        sh.column_dimensions[chr(ord('A')+i-1)].width = 15
        sh.cell(1,i).font = font1
    for i in range(1,sh.max_row+1):
        sh.row_dimensions[i].height = 30
        sh.cell(i,1).font = font1
    for row in sh.iter_rows(min_row=2,min_col=2):
        for cell in row:
            cell.font = font2
    for row in range(1,sh.max_row+1):
        if row % 2 == 0:
            for col in range(1,sh.max_column+1):
                sh.cell(row,col).fill =  PatternFill('solid','d0d0d0')
    for row in sh.iter_rows(min_row=2,min_col=2):
        for cell in row:
            if cell.value >= 130:
                cell.fill = PatternFill('solid','00EC00')
            elif cell.value < 70:
                cell.fill = PatternFill('solid','FF5151')
                cell.font = font3
    wb.save('财务报表.xlsx')
Create_Excel(infos)
Beautify()