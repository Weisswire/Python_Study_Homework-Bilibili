# 1、将movie中的多个工作表的数据合并到一个工作表中起名为all_movie
# 2、将movie表中的数据合并成一个新的Excel文件名为all_movie

from openpyxl import Workbook, load_workbook

movie_infos = {'movie_infos1':
         [
             ['电影', '评分', '上映时间'],
             ['阿凡达', 9.3, 2021],
             ['五尺天涯', 8.9, 2021],
             ['你好，李焕英', 9.5, 2021],
             ['猫和老鼠', 7.9, 2021],
             ['唐人街探案3', 8.8, 2021],
             ['新神榜：哪吒重生', 8.8, 2021]
         ], 'movie_infos2':[
             ['电影', '评分', '上映时间'],
             ['送你一朵小红花', 8.6, 2020],
            ['温暖的抱抱', 9, 2020],
            ['鬼灭之刃剧场版 无限列车篇', 8.6, 2020],
            ['晴雅集', 9.7, 2020],
            ['战狼2', 9.2, 2020],
            ['我和我的家乡', 9.3, 2020]]
}

def Create_Excel(movie_infos: dict):
    wb = Workbook()
    sh1 = wb.active
    sh1.title = 'movie_infos1'
    for info in movie_infos['movie_infos1']:
        sh1.append(info)
    sh2 = wb.create_sheet('movie_infos2')
    for info in movie_infos['movie_infos2']:
        sh2.append(info)
    wb.save('video.xlsx')

def Merge_Sheet():
    wb = load_workbook('video.xlsx')
    sh1 = wb['movie_infos1']
    sh2 = wb['movie_infos2']
    all_movies = []
    for row in sh1.rows:
        tmp_list = []
        for cell in row:
            tmp_list.append(cell.value)
        all_movies.append(tmp_list)
    for row in sh2.rows:
        if row[0].value == '电影':
            continue
        else:
            tmp_list = []
            for cell in row:
                tmp_list.append(cell.value)
            all_movies.append(tmp_list)
    sh3 = wb.create_sheet('all_movie')
    for info in all_movies:
        sh3.append(info)
    wb.save('video.xlsx')

def Merge_Excel():
    wb1 = load_workbook('video.xlsx')
    wb2 = Workbook()
    sh1 = wb1['movie_infos1']
    sh2 = wb1['movie_infos2']
    all_movies_info = []
    for row in sh1.rows:
        tmp_list = []
        for cell in row:
            tmp_list.append(cell.value)
        all_movies_info.append(tmp_list)
    for row in sh2.rows:
        if row[0].value == '电影':
            continue
        tmp_list = []
        for cell in row:
            tmp_list.append(cell.value)
        all_movies_info.append(tmp_list)
    sh3 = wb2.active
    sh3.title = 'all_movies_infos'
    for info in all_movies_info:
        sh3.append(info)
    wb2.save('all_movie.xlsx')

Create_Excel(movie_infos)
Merge_Sheet()
Merge_Excel()