# 1、创建名称为vedio的excel文件，并将自己喜欢的5个番剧保存到文件中，其中包括：
# 番名、主角名称、评分等级、出品国家
# 2、分别统计vedio表格中评分超过9分的番剧有多少，出品国家为国产的有多少并记录在第二个Sheet表中
from openpyxl import Workbook,load_workbook

def create_excel():
    wb = Workbook()
    sh = wb.active
    sh.title = '番剧信息'
    infos = [
        ['番名', '主角名称', '评分等级', '出品国家'],
        ['冰菓', '千反田爱瑠', 9.8, '日本'],
        ['吹响吧！上低音号', '黄前久美子', 9.8, '日本'],
        ['灵笼', '马克', 9.6, '中国'],
        ['BanG Dream! 少女乐团派对！', '户山香澄', 9.5, '日本'],
        ['那年那兔那些事儿', '兔子', 9.7, '中国'],
    ]
    for info in infos:
        sh.append(info)
    wb.save('video.xlsx')

def calculate():
    wb = load_workbook('video.xlsx')
    sh = wb.create_sheet('9分以上番剧个数&国漫个数')
    sh['A1'] = '9分以上番剧个数'
    sh['A2'] = over_nine()
    sh['B1'] = '国漫个数'
    sh['B2'] = from_China()
    wb.save('video.xlsx')

def over_nine() -> int:
    wb = load_workbook('video.xlsx')
    sh = wb['番剧信息']
    # 以下为第一种实现方法：
    # count_rows = 1
    # count = 0
    # for row in sh.rows:
    #     count_rows += 1
    # for i in range(2,count_rows):
    #     if sh[f'C{i}'].value > 9.0:
    #         count += 1
    # 以下为第二种实现方法：
    count = 0
    for row in sh.rows:
        if (type(row[2].value) == float or type(row[2].value) == int) and row[2].value > 9.0:
            count += 1
    return count

def from_China() -> int:
    wb = load_workbook('video.xlsx')
    sh = wb['番剧信息']
    count_rows = 1
    count = 0
    for row in sh.rows:
        count_rows += 1
    for i in range(2,count_rows):
        if sh[f'D{i}'].value == '中国':
            count += 1
    return count

create_excel()
calculate()
