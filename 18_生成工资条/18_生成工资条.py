# 编写脚本，从文件test.xlsx中提取数据
# 根据每一条数据分别生成当月账单信息

from openpyxl import Workbook,load_workbook

def Create_Bill():
    wb = load_workbook('test.xlsx')
    sh = wb.active
    title = None
    for i,row in enumerate(sh.rows):
        tmp_list = []
        for cell in row:
            tmp_list.append(cell.value)
        if i == 0:
            title = tmp_list
        else:
            wb_tmp = Workbook()
            sh_tmp = wb_tmp.active
            sh_tmp.append(title)
            sh_tmp.append(tmp_list)
            wb_tmp.save(f'{tmp_list[0]}_{tmp_list[1]}月账单.xlsx')
Create_Bill()