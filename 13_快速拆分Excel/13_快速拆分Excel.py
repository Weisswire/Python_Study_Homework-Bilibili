# 1、通过Python创建excel数据并把多个数据加入到excel中第一个Sheet中（具体数据见代码部分）
# 2、根据年限将Sheet1表中的数据进行分割，分别将不同年份的数据保存到相应年的Sheet表中

from openpyxl import load_workbook, Workbook

infos = [
        ['电影', '评分', '上映时间'],
        ['阿凡达', 9.3, 2021],
        ['五尺天涯', 8.9, 2021],
        ['你好，李焕英', 9.5, 2021],
        ['猫和老鼠', 7.9, 2021],
        ['唐人街探案3', 8.8, 2021],
        ['新神榜：哪吒重生', 8.8, 2021],
        ['送你一朵小红花', 8.6, 2020],
        ['温暖的抱抱', 9, 2020],
        ['鬼灭之刃剧场版 无限列车篇', 8.6, 2020],
        ['晴雅集', 9.7, 2020],
        ['战狼2', 9.2, 2020],
        ['我和我的家乡', 9.3, 2020],
]

def Create_Excel(infos: list):
    wb = Workbook()
    sh = wb.active
    sh.title = '影视信息'
    for info in infos:
        sh.append(info)
    wb.save("video.xlsx")

def operate():
    statistics(2020)
    statistics(2021)

def statistics(year:int):
    wb = load_workbook('video.xlsx')
    sh = wb['影视信息']
    sh1 = wb.create_sheet(f'{year}年上映的电影')
    sh1['A1'] = '电影'
    sh1['B1'] = '评分'
    sh1['C1'] = '上映时间'
    data = []
    for row in sh.rows:
        if type(row[2].value) == int and row[2].value == year:
            data.append(row)
    for d in data:
        tmp_list = []
        for tmp in d:
            tmp_list.append(tmp.value)
        sh1.append(tmp_list)
    wb.save('video.xlsx')
Create_Excel(infos)
operate()