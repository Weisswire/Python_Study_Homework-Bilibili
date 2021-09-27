# 1、将t1、t2、t3 Excel文件进行合并，并保存名为t_all的Excel文件

from openpyxl import Workbook, load_workbook

movie_infos = {'movie_infos1':
               [
                   ['阿凡达', 9.3, 2021],
                   ['五尺天涯', 8.9, 2021],
                   ['你好，李焕英', 9.5, 2021],
                   ['猫和老鼠', 7.9, 2021],
                   ['唐人街探案3', 8.8, 2021],
                   ['新神榜：哪吒重生', 8.8, 2021]
               ], 'movie_infos2': [
                   ['送你一朵小红花', 8.6, 2020],
                   ['温暖的抱抱', 9, 2020],
                   ['鬼灭之刃剧场版 无限列车篇', 8.6, 2020]
                ],
               'movie_infos3': [
                   ['晴雅集', 9.7, 2020],
                   ['战狼2', 9.2, 2020],
                   ['我和我的家乡', 9.3, 2020]
                ]
               }

def Create_Excel(infos: list, item: int):
    wb = Workbook()
    sh = wb.active
    for info in infos:
        sh.append(info)
    wb.save(f't{item}.xlsx')

def Merge_Excel():
    all = []
    for i in range(3):
        wb = load_workbook(f't{i+1}.xlsx')
        sh = wb.active
        for row in sh.rows:
            tmp_list = []
            for cell in row:
                tmp_list.append(cell.value)
            all.append(tmp_list)
    wb = Workbook()
    sh = wb.active
    for t in all:
        sh.append(t)
    wb.save('t_all.xlsx')

for i in range(3):
    Create_Excel(movie_infos[f'movie_infos{i+1}'], i+1)
Merge_Excel()