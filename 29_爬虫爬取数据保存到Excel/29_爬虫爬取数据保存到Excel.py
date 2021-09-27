# 1. 编写爬虫提取B站番剧排行榜的第一页信息，包含名称、播放量、点赞量，将这些数据保存到excel中
# 2. 编写爬虫爬取B站动画排行榜第一页的信息，包含名称、播放量、点赞量，保存到excel中
from openpyxl import Workbook,load_workbook
from lxml import etree
import re
import requests

def bangumi_rank():
    wb = Workbook()
    sh = wb.active
    sh.title = "番剧排名信息"
    sh.append(["排名","番剧名称","播放量","点赞量"])
    url = "https://www.bilibili.com/v/popular/rank/bangumi"
    header = {'User-Agent':'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/91.0.4472.124 Safari/537.36'}
    resp = requests.get(url,headers = header)
    url_head = "https://api.bilibili.com/pgc/web/season/stat?season_id="
    urls_tile = re.findall("<a href=\"//www.bilibili.com/bangumi/play/ss(.+)\" target=\"_blank\">",resp.text)
    e = etree.HTML(resp.text)
    names = e.xpath("//a[@class=\"title\"]/text()")
    plays = e.xpath("//span[@class=\"data-box\"][1]/text()")
    likes = []
    for i,info in enumerate(urls_tile):
        print(f"正在爬取第{i+1}条番剧排名信息……")
        likes_info = requests.get(url_head + info, headers = header)
        likes.append(likes_info.json().get("result").get("likes"))
    i = 0
    for name,play,like in zip(names,plays,likes):
        i += 1
        sh.append([i,name,play.strip(),like])
    print("番剧排名信息爬取完毕！")
    wb.save("./29_爬虫爬取数据保存到Excel/排名信息.xlsx")

def douga_rank():
    wb = load_workbook("./29_爬虫爬取数据保存到Excel/排名信息.xlsx")
    sh = wb.create_sheet("动画排名信息")
    sh.append(["排名","动画名称","播放量","点赞量"])
    url = "https://www.bilibili.com/v/popular/rank/douga"
    header = {'User-Agent':'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/91.0.4472.124 Safari/537.36'}
    resp = requests.get(url,headers = header)
    urls = re.findall("<a href=\"//(.+)\" target=\"_blank\">",resp.text)
    e = etree.HTML(resp.text)
    plays = e.xpath("//span[@class=\"data-box\"][1]/text()")
    names = e.xpath("//a[@class=\"title\"]/text()")
    likes = []
    for i,url in enumerate(urls):
        print(f"正在爬取第{i+1}条动画排名信息……")
        likes_info = requests.get("https://" + url,headers = header)
        info = re.findall("点赞数(.+)\" class=\"like\">",likes_info.text)
        likes.append(info[0])
    i = 0
    for name,play,like in zip(names,plays,likes):
        i += 1
        sh.append([i,name,play.strip(),like])
    print("动画排名信息爬取完毕！")
    wb.save("./29_爬虫爬取数据保存到Excel/排名信息.xlsx")

if __name__ == "__main__":
    bangumi_rank()
    douga_rank()