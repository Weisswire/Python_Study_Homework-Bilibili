# 1、编写程序实现在文件tmp中根据某指标生成条图
# 2、编写程序实现在文件tmp中根据某指标生成折线图

from openpyxl import Workbook,load_workbook
from openpyxl.chart import LineChart,PieChart,BarChart,Reference
def CreateExcel():
    wb = Workbook()
    sh = wb.active
    rows = [
        ["月份","月平均气温"],
        [1,1],
        [2,3],
        [3,5],
        [4,9],
        [5,15],
        [6,25],
        [7,30],
        [8,35],
        [9,24],
        [10,13],
        [11,4],
        [12,2],
    ]
    for row in rows:
        sh.append(row)
    wb.save("tmp.xlsx")

def CreateBar():
    wb = load_workbook("tmp.xlsx")
    sh = wb.active
    bc = BarChart()
    bc.title = "2021年月平均气温条形图"
    bc.x_axis.title = "月份"
    bc.y_axis.title = "气温/℃"
    cats = Reference(sh,min_col = 1,min_row = 2,max_row = 13)
    data = Reference(sh,min_col = 2,min_row = 1,max_row = 13)
    bc.add_data(data,titles_from_data = True)
    bc.set_categories(cats)
    sh.add_chart(bc,"A16")
    wb.save("tmp.xlsx")

def CreateLine():
    wb = load_workbook("tmp.xlsx")
    sh = wb.active
    lc = LineChart()
    lc.title = "2021年月平均气温折线图"
    lc.x_axis.title = "月份"
    lc.y_axis.title = "气温/℃"
    data = Reference(sh,min_col = 2,min_row = 1,max_row = 13)
    lc.add_data(data,titles_from_data = True)
    sh.add_chart(lc,"A31")
    wb.save("tmp.xlsx")

def CreatePie():
    wb = load_workbook("tmp.xlsx")
    sh = wb.active
    pc = PieChart()
    pc.title = "2021年月平均气温饼状图"
    data = Reference(sh,min_row = 1,min_col = 2,max_row = 13)
    cats = Reference(sh,min_row = 2,min_col = 1,max_row = 13)
    pc.add_data(data,titles_from_data = True)
    pc.set_categories(cats)
    sh.add_chart(pc,"A46")
    wb.save("tmp.xlsx")
CreateExcel()
CreateBar()
CreateLine()
CreatePie()