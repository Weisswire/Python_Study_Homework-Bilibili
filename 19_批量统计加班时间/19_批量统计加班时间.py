# 工作日不得超过1.5小时，节假日不得超过3小时
# 统计每个工作日、节假日超出规定的游戏时间

from openpyxl import Workbook,load_workbook
from openpyxl.styles import Alignment,colors,Font

def Calculate():
    wb = load_workbook('上网时间.xlsx')
    sh = wb.active
    all = []
    for i,row in enumerate(sh.iter_rows(max_col=4)):
        tmp_list = []
        for cell in row:
            tmp_list.append(cell.value)
        if i == 0:
            tmp_list.append('超时时长')
        else:    
            if tmp_list[3] < 6:
                over_time = tmp_list[2] - 90
                if over_time > 0:
                    tmp_list.append(over_time)
                else:
                    tmp_list.append('未超时')
            else:
                over_time = tmp_list[2] - 180
                if over_time > 0:
                    tmp_list.append(over_time)
                else:
                    tmp_list.append('未超时')
        all.append(tmp_list)
    wb = Workbook()
    sh = wb.active
    for a in all:
        sh.append(a)
    wb.save('超时时长.xlsx')

def Beautify():
    font1 = Font(bold=True,color='FF0000')
    font2 = Font(color='00DB00')
    wb = load_workbook('超时时长.xlsx')
    sh = wb.active
    for row in sh.rows:
        for cell in row:
            cell.alignment = Alignment(horizontal='center',vertical='center')
    for row in sh.iter_rows(min_row=2):
        if type(row[4].value) == int and row[4].value > 0:
            row[4].font = font1
        else:
            row[4].font = font2
    wb.save('超时时长.xlsx')

Calculate()
Beautify()