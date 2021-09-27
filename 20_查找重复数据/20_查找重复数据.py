# 1、编写程序实现在文件tmp中找出重复的数据有哪些
# 2、将找出重复的数据进行删除，并把剩余的数据保存到一个新的工作表中

from openpyxl import load_workbook,Workbook

index = []
tmp = []
def Find_Repeat():
    wb = load_workbook('tmp.xlsx')
    sh = wb.active
    for i,row in enumerate(sh.rows):
       if row[5].value in tmp:
           index.append(i)
       else:
           tmp.append(row[5].value) 

def Delete_Copy():
    wb = load_workbook('tmp.xlsx')
    sh = wb.active
    maxRow = sh.max_row
    count = 0
    for i in range(maxRow):
        if i in index:
            sh.delete_rows(i+1-count)
            count += 1
    wb.save('tmp.xlsx')
    wb.save('tmp1.xlsx')
Find_Repeat()
Delete_Copy()