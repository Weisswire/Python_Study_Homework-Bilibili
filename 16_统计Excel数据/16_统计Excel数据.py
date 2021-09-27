# 将z1、z2、z3文件中的数据进行合并并统计每一类播放量的总和并保存到名为z_all的Excel文件

from openpyxl import load_workbook,Workbook

wb = Workbook()
all = []
titles = []
for i in range(1,4):
    wb1 = load_workbook(f'z{i}.xlsx')
    sh = wb1.active
    titles.append(sh.title)
    sum = 0
    # 以下是秀操作版
    for row in sh.iter_rows(min_col=2):
        for cell in row:
            sum += cell.value
    # 以下是原版
    # for row in (1,sh.max_row + 1):
    #     sum += sh.cell(row,sh.max_column).value
    # 以下是无脑版
    # for row in sh.rows:
    #     for cell in row:
    #         tmp = str(cell.value)
    #         if tmp.isdigit():
    #             sum += cell.value
    all.append(sum)
sh = wb.active
sh.title = '汇总'
sh.append(titles)
sh.append(all)
wb.save('z_all.xlsx')